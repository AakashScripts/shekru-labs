from django.shortcuts import render, redirect
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, throttle_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import HttpResponse, JsonResponse
from .models import ExcelData, UploadedExcel
from .serializers import ExcelDataSerializer
import openpyxl
from django.http import HttpResponse
from datetime import datetime
from django.views.generic import ListView
import zipfile
import io
import os
from rest_framework.parsers import MultiPartParser, FormParser
import traceback
from .models import DownloadHistory
from django.core.files.storage import FileSystemStorage
from .tasks import process_excel_file, process_zip_file
import tempfile
import logging
from django.contrib.auth.mixins import LoginRequiredMixin 
import pandas as pd
from django.conf import settings
from .resource import ExcelDataResource
from django.contrib import messages
from tablib import Dataset
from django.core.paginator import Paginator
from django.core.cache import cache
from rest_framework.throttling import AnonRateThrottle
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import ExcelData
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import Q
from .models import ExcelData
import csv
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required

logger = logging.getLogger(__name__)

class ExcelDataListView(ListView):
    model = ExcelData
    template_name = 'aap_api/item_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        queryset = ExcelData.objects.filter(is_visible=True)

        # Apply filters
        name_search = self.request.GET.get('name_search')
        job_title = self.request.GET.get('job_title')
        location = self.request.GET.get('location')
        company = self.request.GET.get('company')

        if name_search:
            queryset = queryset.filter(name__icontains=name_search)
        if job_title:
            queryset = queryset.filter(job_title=job_title)
        if location:
            queryset = queryset.filter(current_location=location)
        if company:
            queryset = queryset.filter(current_company_name=company)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter options
        context['job_titles'] = ExcelData.objects.values_list(
            'job_title', flat=True).distinct().exclude(job_title='')
        context['locations'] = ExcelData.objects.values_list(
            'current_location', flat=True).distinct().exclude(current_location='')
        context['companies'] = ExcelData.objects.values_list(
            'current_company_name', flat=True).distinct().exclude(current_company_name='')
        
        # Add current filter values
        context['current_filters'] = {
            'name_search': self.request.GET.get('name_search', ''),
            'job_title': self.request.GET.get('job_title', ''),
            'location': self.request.GET.get('location', ''),
            'company': self.request.GET.get('company', '')
        }
        
        return context

@method_decorator(csrf_exempt, name='dispatch')
class ExcelDataViewSet(viewsets.ModelViewSet):
    """
    ViewSet for viewing and editing ExcelData instances.
    Provides default `list`, `create`, `retrieve`, `update` and `delete` actions.
    Additional actions:
    - import_excel: POST endpoint to import data from Excel file
    - export_excel: GET endpoint to export data to Excel file
    - import_zip_excel: POST endpoint to import data from ZIP file containing multiple Excel files
    """
    queryset = ExcelData.objects.select_related().all()
    serializer_class = ExcelDataSerializer
    permission_classes = [AllowAny]
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = []  # Remove authentication for testing
    page_size = 50  # Add pagination

    def get_queryset(self):
        queryset = ExcelData.objects.filter(is_visible=True).exclude(
            name='',
            job_title='',
            email_id='',
            phone_number='',
            current_location='',
            total_experience='',
            current_company_name=''
        )
        return queryset.order_by('-created_at')

    def get_column_indices(self, worksheet):
        """Helper function to find column indices based on headers"""
        headers = {}
        header_mappings = {
            # Standard format
            'job_title': ['job_title', 'job title'],
            'date_of_application': ['date_of_application', 'date of application'],
            'email_id': ['email_id', 'email id'],
            'phone_number': ['phone_number', 'phone number'],
            'current_location': ['current_location', 'current location'],
            'preferred_locations': ['preferred_locations', 'preferred locations'],
            'total_experience': ['total_experience', 'total experience'],
            'current_company_name': ['current_company_name', 'current company name', 'curr. company name'],
            'current_company_designation': ['current_company_designation', 'current company designation', 'curr. company designation'],
            'key_skills': ['key_skills', 'key skills'],
            'annual_salary': ['annual_salary', 'annual salary'],
            'notice_period_availability_to_join': ['notice_period_availability_to_join', 'notice period availability to join', 'notice period/ availability to join'],
            'resume_headline': ['resume_headline', 'resume headline'],
            'under_graduation_degree': ['under_graduation_degree', 'under graduation degree'],
            'ug_specialization': ['ug_specialization', 'ug specialization'],
            'ug_university_institute_name': ['ug_university_institute_name', 'ug university/institute name'],
            'ug_graduation_year': ['ug_graduation_year', 'ug graduation year'],
            'post_graduation_degree': ['post_graduation_degree', 'post graduation degree'],
            'pg_specialization': ['pg_specialization', 'pg specialization'],
            'pg_university_institute_name': ['pg_university_institute_name', 'pg university/institute name'],
            'pg_graduation_year': ['pg_graduation_year', 'pg graduation year'],
            'doctorate_degree': ['doctorate_degree', 'doctorate degree'],
            'doctorate_specialization': ['doctorate_specialization', 'doctorate specialization']
        }

        for cell in worksheet[1]:  # First row
            if cell.value:
                header = str(cell.value).lower().strip()
                # Find the matching standard header
                for standard_header, variations in header_mappings.items():
                    if header in variations:
                        headers[standard_header] = cell.column - 1
                        break
                else:
                    # If no mapping found, use the header as is
                    headers[header] = cell.column - 1
        return headers

    def parse_date(self, value):
        """Helper function to parse dates from Excel"""
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            # Try different date formats
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(str(value).split()[0], fmt).date()
                except ValueError:
                    continue
            return None
        except Exception:
            return None

    @action(detail=False, methods=['POST'])
    def import_excel(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=400)

        excel_file = request.FILES['file']
        try:
            df = pd.read_excel(excel_file)
            print(f"Reading {len(df)} rows from Excel")  # Debug print
            
            for _, row in df.iterrows():
                record = ExcelData.objects.create(
                    job_title=row.get('job_title', ''),
                    date_of_application=row.get('date_of_application'),
                    name=row.get('name', ''),
                    email_id=row.get('email_id', ''),
                    phone_number=str(row.get('phone_number', '')),
                    current_location=row.get('current_location', ''),
                    total_experience=str(row.get('total_experience', '')),
                    current_company_name=row.get('current_company_name', '')
                )
                print(f"Created record: {record.id}")  # Debug print
            
            return Response({
                'message': f'Successfully imported {len(df)} records',
                'rows_imported': len(df)
            })
        except Exception as e:
            print(f"Error importing Excel: {str(e)}")  # Debug print
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['GET'])
    def export_excel(self, request):
        try:
            queryset = self.get_queryset()
            
            # Convert queryset to DataFrame
            data = []
            for item in queryset:
                data.append({
                    'Name': item.name,
                    'Job Title': item.job_title,
                    'Email': item.email_id,
                    'Phone': item.phone_number,
                    'Location': item.current_location,
                    'Experience': item.total_experience,
                    'Company': item.current_company_name,
                    'Date Added': item.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            df = pd.DataFrame(data)
            
            # Create response
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename=data_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            # Write to excel
            df.to_excel(response, index=False, engine='openpyxl')
            return response
            
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=500)

    @action(detail=False, methods=['POST'])
    def import_zip_excel(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        zip_file = request.FILES['file']
        if not zip_file.name.endswith('.zip'):
            return JsonResponse({'error': 'File must be a ZIP archive'}, status=400)

        try:
            # Create a temporary directory to extract files
            with tempfile.TemporaryDirectory() as temp_dir:
                # Save the uploaded zip file
                temp_zip_path = os.path.join(temp_dir, 'upload.zip')
                with open(temp_zip_path, 'wb') as f:
                    for chunk in zip_file.chunks():
                        f.write(chunk)

                # Extract the zip file
                with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)

                # Process each Excel file in the extracted directory
                successful_imports = 0
                failed_imports = 0

                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.endswith(('.xlsx', '.xls')):
                            file_path = os.path.join(root, file)
                            try:
                                # Read Excel file
                                df = pd.read_excel(file_path)
                                
                                # Process each row
                                for _, row in df.iterrows():
                                    ExcelData.objects.create(
                                        job_title=row.get('job_title', ''),
                                        date_of_application=row.get('date_of_application'),
                                        name=row.get('name', ''),
                                        email_id=row.get('email_id', ''),
                                        phone_number=row.get('phone_number', ''),
                                        current_location=row.get('current_location', ''),
                                        preferred_locations=row.get('preferred_locations', ''),
                                        total_experience=row.get('total_experience', ''),
                                        current_company_name=row.get('current_company_name', '')
                                    )
                                successful_imports += 1
                            except Exception as e:
                                failed_imports += 1
                                print(f"Error processing {file}: {str(e)}")

                return JsonResponse({
                    'message': 'ZIP file processed successfully',
                    'successful_imports': successful_imports,
                    'failed_imports': failed_imports
                })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    @action(detail=False, methods=['GET'])
    def task_status(self, request):
        """
        Check the status of a running import task
        """
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'No task ID provided'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            task = process_zip_file.AsyncResult(task_id)
            
            if task.state == 'PENDING':
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task is pending...',
                        'progress': 0,
                        'current_file': None,
                        'total_processed': 0
                    }
                }
            elif task.state == 'PROGRESS':
                info = task.info or {}
                response = {
                    'state': task.state,
                    'status': {
                        'progress': info.get('progress', 0),
                        'current_file': info.get('current_file', ''),
                        'total_processed': info.get('total_processed', 0),
                        'total_files': info.get('total_files', 0),
                        'processed_files': info.get('processed_files', 0)
                    }
                }
            elif task.state == 'SUCCESS':
                result = task.get()  # Get the actual result
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task completed successfully',
                        'total_processed': result.get('total_processed', 0),
                        'total_files': result.get('total_files', 0),
                        'successful_files': result.get('successful_files', 0),
                        'failed_files': result.get('failed_files', 0),
                        'results': result.get('results', [])
                    }
                }
            elif task.state == 'FAILURE':
                response = {
                    'state': task.state,
                    'status': {
                        'message': 'Task failed',
                        'error': str(task.info) if task.info else 'Unknown error occurred'
                    }
                }
            else:
                response = {
                    'state': task.state,
                    'status': {
                        'message': f'Task is in {task.state} state',
                        'info': str(task.info) if task.info else None
                    }
                }
            
            return Response(response)
            
        except Exception as e:
            logger.error(f"Error checking task status: {str(e)}")
            return Response({
                'state': 'ERROR',
                'status': {
                    'message': 'Error checking task status',
                    'error': str(e)
                }
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'count': queryset.count(),
            'data': serializer.data
        })

    @action(detail=False, methods=['GET'])
    def test_connection(self, request):
        return JsonResponse({
            'status': 'success',
            'message': 'API is working'
        })

    @action(detail=False, methods=['get'])
    def all(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data,
            'total_count': queryset.count()
        })

class DownloadHistoryListView(ListView):
    model = DownloadHistory
    template_name = 'aap_api/download_history.html'
    context_object_name = 'download_history'
    paginate_by = 20  # Adjust as needed

    def get_queryset(self):
        return DownloadHistory.objects.order_by('-download_time')

def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_extension = os.path.splitext(file.name)[1].lower()

        try:
            if file_extension == '.zip':
                # Handle ZIP file
                zip_data = io.BytesIO(file.read())
                files_data = []

                with zipfile.ZipFile(zip_data) as zip_archive:
                    excel_files = [f for f in zip_archive.namelist() if f.lower().endswith(('.xlsx', '.xls'))]

                    for excel_file in excel_files:
                        excel_data = zip_archive.read(excel_file)
                        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
                        ws = wb.active
                        data = []

                        headers = [cell.value for cell in ws[1]]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            row_data = dict(zip(headers, row))
                            data.append(row_data)

                        files_data.append({
                            'name': excel_file,
                            'data': data[:100]  # Limit preview to 100 rows
                        })

                return JsonResponse({'files': files_data})

            elif file_extension in ['.xlsx', '.xls']:
                # Handle Excel file
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                data = []

                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    data.append(row_data)

                return JsonResponse({
                    'files': [{
                        'name': file.name,
                        'data': data[:100]  # Limit preview to 100 rows
                    }]
                })
            else:
                return JsonResponse({'error': 'Invalid file format'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return render(request, 'aap_api/upload.html')

class CustomAnonRateThrottle(AnonRateThrottle):
    rate = '50/minute'  # Adjust this value based on your needs

@api_view(['GET'])
@throttle_classes([CustomAnonRateThrottle])
def get_all_data(request):
    try:
        logger.debug(f"Request received with params: {request.GET}")
        offset = int(request.GET.get('offset', 0))
        limit = int(request.GET.get('limit', 2000))
        
        logger.debug(f"Fetching data with offset={offset}, limit={limit}")
        total_count = ExcelData.objects.count()
        logger.debug(f"Total records in database: {total_count}")
        
        records = ExcelData.objects.all()\
            .order_by('id')\
            .values(
                'id', 'name', 'job_title', 'email_id',
                'phone_number', 'current_location',
                'total_experience', 'current_company_name'
            )[offset:offset + limit]
        
        response_data = {
            'status': 'success',
            'total_count': total_count,
            'offset': offset,
            'limit': limit,
            'data': list(records)
        }
        logger.debug(f"Returning {len(records)} records")
        
        return Response(response_data)
        
    except Exception as e:
        logger.error(f"Error in get_all_data: {str(e)}", exc_info=True)
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def item_list(request):
    # Get filter parameters
    name_search = request.GET.get('name_search', '')
    job_title = request.GET.get('job_title', '')
    location = request.GET.get('location', '')
    company = request.GET.get('company', '')

    # Base queryset
    queryset = ExcelData.objects.all()

    # Apply filters
    if name_search:
        queryset = queryset.filter(name__icontains(name_search))
    if job_title:
        queryset = queryset.filter(job_title=job_title)
    if location:
        queryset = queryset.filter(current_location=location)
    if company:
        queryset = queryset.filter(current_company_name=company)

    # Get unique values for dropdowns
    job_titles = ExcelData.objects.values_list('job_title', flat=True).distinct()
    locations = ExcelData.objects.values_list('current_location', flat=True).distinct()
    companies = ExcelData.objects.values_list('current_company_name', flat=True).distinct()

    # Pagination
    paginator = Paginator(queryset, 30)  # Show 10 items per page
    page = request.GET.get('page')
    items = paginator.get_page(page)

    context = {
        'items': items,
        'total_records': queryset.count(),
        'job_titles': [(jt, jt) for jt in job_titles if jt],
        'locations': [loc for loc in locations if loc],
        'companies': [comp for comp in companies if comp],
        'current_filters': {
            'name_search': name_search,
            'job_title': job_title,
            'location': location,
            'company': company,
        },
        'is_paginated': items.has_other_pages(),
        'page_obj': items,
    }

    return render(request, 'aap_api/item_list.html', context)


@api_view(['POST'])
def delete_null_records(request):
    try:
        initial_count = ExcelData.objects.filter(is_active=True).count()
        
        # Find records with null/empty values
        null_records = ExcelData.objects.filter(
            Q(name__isnull=True) | Q(name='') |
            Q(job_title__isnull=True) | Q(job_title='') |
            Q(email_id__isnull=True) | Q(email_id='') |
            Q(phone_number__isnull=True) | Q(phone_number='') |
            Q(current_location__isnull=True) | Q(current_location='') |
            Q(total_experience__isnull=True) | Q(total_experience='')
        )
        
        # Soft delete instead of actual deletion
        for record in null_records:
            record.soft_delete()

        final_count = ExcelData.objects.filter(is_active=True).count()
        
        return Response({
            'status': 'success',
            'message': f'Successfully hidden {initial_count - final_count} null records',
            'hidden_count': initial_count - final_count,
            'remaining_count': final_count
        })
    except Exception as e:
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=500)

@api_view(['GET'])
def get_all_data(request):
    try:
        offset = int(request.GET.get('offset', 0))
        limit = int(request.GET.get('limit', 2000))
        
        # Only get visible records
        queryset = ExcelData.objects.filter(is_visible=True)
        total_count = queryset.count()
        
        records = queryset.order_by('id')[offset:offset + limit].values()
        
        return Response({
            'status': 'success',
            'total_count': total_count,
            'data': list(records)
        })
    except Exception as e:
        return Response({
            'status': 'error',
            'message': str(e)
        }, status=500)

from django.http import JsonResponse
from .models import ExcelData
from .serializers import ExcelDataSerializer

def get_all_data(request):
    try:
        # Get all visible records
        records = ExcelData.objects.filter(is_visible=True)
        
        # Serialize the data
        serializer = ExcelDataSerializer(records, many=True)
        
        # Return JSON response
        return JsonResponse({
            'status': 'success',
            'count': records.count(),
            'data': serializer.data
        }, safe=False)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def index(request):
    # Get filter parameters
    name_search = request.GET.get('name_search', '').strip()
    job_title = request.GET.get('job_title', '').strip()
    location = request.GET.get('location', '').strip()
    company = request.GET.get('company', '').strip()

    # Base queryset
    queryset = ExcelData.objects.filter(is_visible=True)

    # Apply filters
    if name_search:
        queryset = queryset.filter(
            Q(name__icontains(name_search)) |
            Q(email_id__icontains(name_search))
        )
    if job_title:
        queryset = queryset.filter(job_title__iexact=job_title)
    if location:
        queryset = queryset.filter(current_location__iexact=location)
    if company:
        queryset = queryset.filter(current_company_name__iexact=company)

    # Order queryset
    queryset = queryset.order_by('-created_at')

    # Pagination
    paginator = Paginator(queryset, 30)  # Show 30 records per page
    page = request.GET.get('page')
    try:
        items = paginator.page(page)
    except PageNotAnInteger:
        items = paginator.page(1)
    except EmptyPage:
        items = paginator.page(paginator.num_pages)

    # Get unique values for dropdowns
    job_titles = list(dict.fromkeys(
        ExcelData.objects.exclude(job_title='')
        .values_list('job_title', flat=True)
        .distinct()
        .order_by('job_title')
    ))

    locations = list(dict.fromkeys(
        ExcelData.objects.exclude(current_location='')
        .values_list('current_location', flat=True)
        .distinct()
        .order_by('current_location')
    ))

    companies = list(dict.fromkeys(
        ExcelData.objects.exclude(current_company_name='')
        .values_list('current_company_name', flat=True)
        .distinct()
        .order_by('current_company_name')
    ))

    total_records = ExcelData.objects.filter(is_visible=True).count()
    filtered_records = queryset.count()

    context = {
        'items': items,
        'job_titles': job_titles,
        'locations': locations,
        'companies': companies,
        'current_filters': {
            'name_search': name_search,
            'job_title': job_title,
            'location': location,
            'company': company
        },
        'total_records': total_records,
        'filtered_records': filtered_records,
        'current_page': page
    }

    return render(request, 'aap_api/index.html', context)

@require_POST
def download_selected(request):
    selected_ids = request.POST.get('selected_ids', '').split(',')
    records = ExcelData.objects.filter(id__in=selected_ids)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="selected_records.csv"'
    
    writer = csv.writer(response)
    # Write headers
    writer.writerow([
        'Name', 'Job Title', 'Email', 'Phone Number',
        'Location', 'Experience', 'Company'
    ])
    
    # Write data
    for record in records:
        writer.writerow([
            record.name,
            record.job_title,
            record.email_id,
            record.phone_number,
            record.current_location,
            record.total_experience,
            record.current_company_name
        ])
    
    return response

def login_view(request):
    if request.user.is_authenticated:
        return redirect('aap_api:index')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('aap_api:index')
        else:
            messages.error(request, 'Invalid credentials')
    
    return render(request, 'aap_api/item_list.html')



